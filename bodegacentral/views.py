from functools import wraps

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_GET, require_POST
from openpyxl.styles import Font

from bodegacentral.forms import CentralForm
from pedidos.models import UserModulePermission

from .models import Central


# Esta funcion valida si un usuario tiene permiso para una accion especifica del modulo.
def _user_has_module_permission(user, module, action):
    if user.is_superuser:
        return True

    permiso = UserModulePermission.objects.filter(user=user, module=module).first()
    if not permiso:
        return False

    permisos = {
        'view': permiso.can_view,
        'create': permiso.can_create,
        'edit': permiso.can_edit,
        'delete': permiso.can_delete,
        'export': permiso.can_export,
        'report': permiso.can_report,
    }
    return permisos.get(action, False)


# Esta funcion crea un decorador para proteger vistas por modulo y tipo de accion.
def require_module_permission(module, action):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if _user_has_module_permission(request.user, module, action):
                return view_func(request, *args, **kwargs)

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'ok': False, 'message': 'No tienes permisos para esta acción.'}, status=403)

            messages.error(request, 'No tienes permisos para esta acción en este módulo.')
            return redirect('inicio')

        return _wrapped_view

    return decorator


# Esta funcion muestra el listado principal de productos en bodega central.
@login_required
@require_module_permission('bodegacentral', 'view')
def index(request):
    centrales = Central.objects.all()
    return render(request, 'central.html', context={'centrales': centrales})


# Esta funcion procesa y muestra el formulario basico de registro de productos.
def formulario(request):
    if request.method == 'POST':
        form = CentralForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/bodegacentral')
    else:
        form = CentralForm()

    return render(request, 'central_form.html', {'form': form})


# Esta funcion normaliza valores vacios o nulos para exportarlos al archivo Excel.
def limpiar_valor(valor):
    return '' if valor in (None, '', 0) else valor


# Esta funcion exporta el inventario de bodega central a un archivo Excel.
@login_required
@require_module_permission('bodegacentral', 'export')
def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    columnas = [
        'Categoria', 'Empresa', 'Ubicacion', 'Cod_Ean', 'Cod_Dun', 'Cod_Sistema',
        'Descripcion', 'Unidad', 'Pack', 'FactorX', 'Cajas', 'Saldo',
        'Stock_Fisico', 'Observacion', 'Fecha_Venc', 'Fecha_Imp',
        'Contenedor', 'Fecha_Inv', 'Encargado',
    ]

    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    data = Central.objects.all()

    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=limpiar_valor(item.categoria))
        ws.cell(row=row_num, column=2, value=limpiar_valor(item.empresa))
        ws.cell(row=row_num, column=3, value=limpiar_valor(item.ubicacion))
        ws.cell(row=row_num, column=4, value=limpiar_valor(item.cod_ean))
        ws.cell(row=row_num, column=5, value=limpiar_valor(item.cod_dun))
        ws.cell(row=row_num, column=6, value=limpiar_valor(item.cod_sistema))
        ws.cell(row=row_num, column=7, value=limpiar_valor(item.descripcion))
        ws.cell(row=row_num, column=8, value=limpiar_valor(item.unidad))
        ws.cell(row=row_num, column=9, value=limpiar_valor(item.pack))
        ws.cell(row=row_num, column=10, value=limpiar_valor(item.factorx))
        ws.cell(row=row_num, column=11, value=limpiar_valor(item.cajas))
        ws.cell(row=row_num, column=12, value=limpiar_valor(item.saldo))
        ws.cell(row=row_num, column=13, value=limpiar_valor(item.stock_fisico))
        ws.cell(row=row_num, column=14, value=limpiar_valor(item.observacion))
        ws.cell(row=row_num, column=15, value=limpiar_valor(item.fecha_venc))
        ws.cell(row=row_num, column=16, value=limpiar_valor(item.fecha_imp))
        ws.cell(row=row_num, column=18, value=limpiar_valor(item.fecha_inv))
        ws.cell(row=row_num, column=19, value=limpiar_valor(item.encargado))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventario_Bodega_Lingues.xlsx"'

    wb.save(response)
    return response


# Esta funcion permite editar un registro existente de bodega central.
@login_required
@require_module_permission('bodegacentral', 'edit')
def editar_central(request, id):
    central = get_object_or_404(Central, id=id)

    if request.method == 'POST':
        form = CentralForm(request.POST, instance=central)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = CentralForm(instance=central)

    return render(request, 'editar_central.html', {'form': form, 'central': central})


# Esta funcion confirma y ejecuta la eliminacion de un registro de bodega central.
@login_required
@require_module_permission('bodegacentral', 'delete')
def eliminar_central(request, id):
    central = get_object_or_404(Central, id=id)

    if request.method == 'POST':
        central.delete()
        return redirect('index')

    return render(request, 'eliminar_central.html', {'central': central})


# Esta funcion gestiona el formulario de creacion o edicion con autocompletado por producto.
@login_required
@require_module_permission('bodegacentral', 'create')
def formulario_central(request, id=None):
    mensaje = ''
    if id:
        instancia = Central.objects.get(id=id)
    else:
        instancia = None

    if request.method == 'POST':
        form = CentralForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            return redirect('/bodegacentral')
            mensaje = 'Producto guardado correctamente.'
    else:
        form = CentralForm(instance=instancia)

    return render(request, 'central_form.html', {'form': form, 'mensaje': mensaje})


# Esta funcion busca productos por coincidencia de codigo DUN para autocompletado.
@login_required
@require_module_permission('bodegacentral', 'view')
def buscar_producto(request):
    cod_dun = request.GET.get('cod_dun', '').strip()
    resultados = []

    if cod_dun:
        productos = Central.objects.filter(cod_dun__icontains=cod_dun)[:10]
        for p in productos:
            resultados.append({
                'cod_dun': p.cod_dun,
                'cod_ean': p.cod_ean,
                'cod_sistema': p.cod_sistema,
                'descripcion': p.descripcion,
            })

    return JsonResponse({'resultados': resultados})


# Esta funcion muestra la vista de cambio de ubicacion con cookie CSRF asegurada.
@login_required
@require_module_permission('bodegacentral', 'view')
@ensure_csrf_cookie
def cambio_ubicacion(request):
    return render(request, 'bodegacentral/cambio_ubicacion.html')


# Esta funcion lista productos que pertenecen a una ubicacion especifica.
@login_required
@require_module_permission('bodegacentral', 'view')
@require_GET
def buscar_por_ubicacion(request):
    ubicacion = request.GET.get('ubicacion', '').strip()
    if not ubicacion:
        return JsonResponse({'ok': False, 'message': 'Debes ingresar una ubicacion.'}, status=400)

    productos = Central.objects.filter(ubicacion__iexact=ubicacion)
    if not productos.exists():
        return JsonResponse({'ok': False, 'message': 'No se encontraron productos en esa ubicacion.'}, status=404)

    lista_productos = []
    for p in productos:
        lista_productos.append({
            'id': p.id,
            'cod_dun': p.cod_dun or '',
            'cod_ean': p.cod_ean or '',
            'cod_sistema': p.cod_sistema or '',
            'descripcion': p.descripcion or '',
            'cajas': p.cajas if p.cajas is not None else 0,
            'ubicacion': p.ubicacion or '',
        })

    return JsonResponse({
        'ok': True,
        'productos': lista_productos,
    })


# Esta funcion mueve cajas entre ubicaciones conservando consistencia en base de datos.
@login_required
@require_module_permission('bodegacentral', 'edit')
@require_POST
def mover_producto(request):
    producto_id = request.POST.get('producto_id', '').strip()
    cantidad_mover = request.POST.get('cantidad_mover', '').strip()
    nueva_ubicacion = request.POST.get('nueva_ubicacion', '').strip()

    if not producto_id:
        return JsonResponse({'ok': False, 'message': 'Producto invalido.'}, status=400)

    if not cantidad_mover or not cantidad_mover.isdigit():
        return JsonResponse({'ok': False, 'message': 'Cantidad invalida.'}, status=400)

    cantidad_mover = int(cantidad_mover)
    if cantidad_mover <= 0:
        return JsonResponse({'ok': False, 'message': 'La cantidad debe ser mayor a 0.'}, status=400)

    if not nueva_ubicacion:
        return JsonResponse({'ok': False, 'message': 'Debes ingresar la nueva ubicacion.'}, status=400)

    producto_origen = Central.objects.filter(id=producto_id).first()
    if not producto_origen:
        return JsonResponse({'ok': False, 'message': 'Producto no encontrado.'}, status=404)

    if producto_origen.cajas is None or producto_origen.cajas < cantidad_mover:
        return JsonResponse({'ok': False, 'message': f'No hay suficientes cajas. Disponible: {producto_origen.cajas or 0}'}, status=400)

    try:
        with transaction.atomic():
            producto_origen.cajas -= cantidad_mover
            if producto_origen.stock_fisico:
                producto_origen.stock_fisico = producto_origen.cajas * (producto_origen.factorx or 1)

            if producto_origen.cajas == 0:
                producto_origen.delete()
            else:
                producto_origen.save()

            producto_destino = Central.objects.filter(
                cod_sistema=producto_origen.cod_sistema,
                ubicacion__iexact=nueva_ubicacion,
            ).first()

            if producto_destino:
                producto_destino.cajas = (producto_destino.cajas or 0) + cantidad_mover
                if producto_destino.stock_fisico is not None:
                    producto_destino.stock_fisico = producto_destino.cajas * (producto_destino.factorx or 1)
                producto_destino.save()
            else:
                Central.objects.create(
                    categoria=producto_origen.categoria,
                    empresa=producto_origen.empresa,
                    ubicacion=nueva_ubicacion,
                    cod_ean=producto_origen.cod_ean,
                    cod_dun=producto_origen.cod_dun,
                    cod_sistema=producto_origen.cod_sistema,
                    descripcion=producto_origen.descripcion,
                    unidad=producto_origen.unidad,
                    pack=producto_origen.pack,
                    factorx=producto_origen.factorx,
                    cajas=cantidad_mover,
                    saldo=producto_origen.saldo,
                    stock_fisico=cantidad_mover * (producto_origen.factorx or 1) if producto_origen.factorx else None,
                    observacion=producto_origen.observacion,
                    fecha_inv=producto_origen.fecha_inv,
                    encargado=producto_origen.encargado,
                    fecha_venc=producto_origen.fecha_venc,
                    fecha_imp=producto_origen.fecha_imp,
                    numero_contenedor=producto_origen.numero_contenedor,
                )

        return JsonResponse({
            'ok': True,
            'message': f'Se movieron {cantidad_mover} cajas a {nueva_ubicacion} correctamente.',
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'message': f'Error al mover producto: {str(e)}'}, status=500)


# Esta funcion muestra un resumen de inventario para validacion operativa.
@login_required
@require_module_permission('bodegacentral', 'view')
def resumen_central(request):
    datos = Central.objects.all().values(
        'id', 'cod_dun', 'cod_ean', 'cod_sistema', 'descripcion', 'cajas', 'stock_fisico', 'ubicacion',
    )
    return render(request, 'resumen.html', {'datos': datos})


# Esta funcion confirma y guarda el total fisico informado desde el resumen.
@login_required
@require_module_permission('bodegacentral', 'edit')
@require_POST
def confirmar_resumen_central(request):
    item_id = (request.POST.get('id') or '').strip()
    total_fisico_raw = (request.POST.get('total_fisico') or '').strip()

    if not item_id.isdigit():
        return JsonResponse({'ok': False, 'message': 'ID invalido.'}, status=400)

    if total_fisico_raw == '' or not total_fisico_raw.lstrip('-').isdigit():
        return JsonResponse({'ok': False, 'message': 'total_fisico debe ser un numero entero.'}, status=400)

    total_fisico = int(total_fisico_raw)
    if total_fisico < 0:
        return JsonResponse({'ok': False, 'message': 'total_fisico no puede ser negativo.'}, status=400)

    producto = Central.objects.filter(id=int(item_id)).first()
    if not producto:
        return JsonResponse({'ok': False, 'message': 'Registro no encontrado.'}, status=404)

    cajas_actuales = producto.cajas or 0
    actualizo_bd = False

    if cajas_actuales != total_fisico:
        producto.cajas = total_fisico
        producto.save(update_fields=['cajas'])
        actualizo_bd = True

    confirmados = request.session.get('resumen_central_confirmados', {})
    confirmados[str(producto.id)] = {
        'cajas_confirmadas': total_fisico,
    }
    request.session['resumen_central_confirmados'] = confirmados

    return JsonResponse({
        'ok': True,
        'updated': actualizo_bd,
        'cajas': total_fisico,
        'message': 'Registro confirmado.',
    })
