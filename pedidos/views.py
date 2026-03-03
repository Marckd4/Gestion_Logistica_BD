from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from io import BytesIO
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models.functions import Replace, Upper
from django.db.models import Value
from django.db.models import Q
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.db import transaction
from decimal import Decimal, ROUND_HALF_UP
from decimal import InvalidOperation

from .forms import NotaVentaForm, DetalleNotaFormSet
from .models import NotaVenta, DetalleNota
from bodegabsf.models import Bsf
from bodegacentral.models import Central


@login_required
def inicio_pedidos(request):
    return redirect('crear_nota')


def _normalizar_rut(rut):
    if not rut:
        return ""
    return rut.replace(".", "").replace("-", "").strip().upper()


def _json_no_cache(payload):
    response = JsonResponse(payload)
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


@login_required
@require_GET
def buscar_cliente_por_rut(request):
    rut = request.GET.get("rut", "")
    nombre = request.GET.get("nombre", "").strip()
    rut_normalizado = _normalizar_rut(rut)

    if not rut_normalizado and not nombre:
        return _json_no_cache({"found": False})

    notas = NotaVenta.objects.all()

    if rut_normalizado:
        notas = notas.exclude(rut_cliente__isnull=True).exclude(rut_cliente="").annotate(
            rut_normalizado=Upper(
                Replace(
                    Replace("rut_cliente", Value("."), Value("")),
                    Value("-"),
                    Value(""),
                )
            )
        ).filter(rut_normalizado=rut_normalizado)

    if nombre:
        notas = notas.exclude(cliente__isnull=True).exclude(cliente="").filter(
            Q(cliente__iexact=nombre) | Q(cliente__icontains=nombre)
        )

    nota = notas.order_by("-fecha").first()

    if not nota:
        return _json_no_cache({"found": False})

    return _json_no_cache(
        {
            "found": True,
            "cliente": {
                "cliente": nota.cliente or "",
                "rut_cliente": nota.rut_cliente or "",
                "giro": nota.giro or "",
                "telefono": nota.telefono or "",
                "direccion": nota.direccion or "",
                "comuna": nota.comuna or "",
                "ciudad": nota.ciudad or "",
                "lugar_de_entrega": nota.lugar_de_entrega or "",
                "persona_responsable": nota.persona_responsable or "",
            },
        }
    )


# ==============================
# CREAR NOTA
# ==============================

@login_required
def crear_nota(request):
    if request.method == 'POST':
        form = NotaVentaForm(request.POST)
        if form.is_valid():
            nota = form.save(commit=False)
            nota.vendedor = request.user
            nota.save()
            messages.success(request, "Nota creada correctamente.")
            return redirect('ingresar_producto', nota_id=nota.id)
    else:
        form = NotaVentaForm()

    return render(request, 'pedidos/crear_nota.html', {'form': form})


@login_required
def ingresar_producto(request, nota_id):
    nota = get_object_or_404(NotaVenta, id=nota_id)
    detalles = nota.detalles.all().order_by('id')
    return render(
        request,
        'pedidos/ingresar_producto.html',
        {
            'nota': nota,
            'detalles': detalles,
            'next_correlativo': detalles.count() + 1,
        }
    )


@login_required
@require_POST
def agregar_detalle_nota(request, nota_id):
    nota = get_object_or_404(NotaVenta, id=nota_id)

    codigo = request.POST.get('cod_sistema', '').strip()
    descripcion_form = request.POST.get('descripcion', '').strip()
    cantidad_raw = request.POST.get('cantidad', '').strip()
    valor_raw = request.POST.get('valor', '').strip()

    if not codigo:
        return _json_no_cache({'ok': False, 'message': 'Código de sistema requerido.'})

    try:
        cantidad = int(cantidad_raw)
        if cantidad <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return _json_no_cache({'ok': False, 'message': 'Cantidad inválida.'})

    try:
        valor = Decimal(valor_raw)
        if valor < 0:
            raise InvalidOperation
    except (TypeError, InvalidOperation, ValueError):
        return _json_no_cache({'ok': False, 'message': 'Valor inválido.'})

    producto_bsf = Bsf.objects.filter(cod_sistema__iexact=codigo).first()
    producto_central = Central.objects.filter(cod_sistema__iexact=codigo).first()
    producto = producto_bsf or producto_central

    descripcion = descripcion_form
    ean = ''
    dun = ''
    formato_venta = ''
    capacidad_x_caja = None

    if producto:
        descripcion = producto.descripcion or descripcion_form
        ean = producto.cod_ean or ''
        dun = producto.cod_dun or ''
        formato_venta = producto.unidad or ''
        capacidad_x_caja = producto.pack if producto.pack else None

    if not descripcion:
        return _json_no_cache({'ok': False, 'message': 'Descripción requerida.'})

    detalle = DetalleNota.objects.create(
        nota=nota,
        codigo=codigo,
        descripcion=descripcion,
        ean=ean,
        dun=dun,
        formato_venta=formato_venta,
        cantidad_solicitada=cantidad,
        capacidad_x_caja=capacidad_x_caja,
        precio_x_caja=valor,
        precio_unitario=valor,
    )

    return _json_no_cache(
        {
            'ok': True,
            'message': 'Producto guardado correctamente.',
            'detalle': {
                'codigo': detalle.codigo or '',
                'descripcion': detalle.descripcion or '',
                'cantidad': detalle.cantidad_solicitada,
                'valor': str(detalle.precio_unitario),
            },
        }
    )


@login_required
@require_POST
def terminar_pedido(request, nota_id):
    nota = get_object_or_404(NotaVenta, id=nota_id)
    nota.estado = 'finalizada'
    nota.save(update_fields=['estado'])
    messages.success(request, f"Pedido de la Nota #{nota.id} finalizado correctamente.")
    return redirect('status_pedido')


def _obtener_ids_seleccionados(request):
    ids_limpios = []
    for nota_id in request.POST.getlist('nota_ids'):
        try:
            valor = int(nota_id)
            if valor > 0 and valor not in ids_limpios:
                ids_limpios.append(valor)
        except (TypeError, ValueError):
            continue
    return ids_limpios


def _exportar_notas_excel(notas):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas de Venta"

    columnas = [
        "Numero Nota",
        "Cliente",
        "Usuario",
        "Fecha",
        "Estado",
        "Tipo Bodega",
        "Neto",
        "IVA",
        "Total",
    ]

    for col_num, titulo in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=col_num, value=titulo)
        celda.font = Font(bold=True)

    for row_num, nota in enumerate(notas, 2):
        ws.cell(row=row_num, column=1, value=nota.id)
        ws.cell(row=row_num, column=2, value=nota.cliente or "")
        ws.cell(row=row_num, column=3, value=nota.vendedor.username if nota.vendedor else "")
        ws.cell(row=row_num, column=4, value=nota.fecha.strftime('%d/%m/%Y %H:%M') if nota.fecha else "")
        ws.cell(row=row_num, column=5, value=nota.get_estado_display())
        ws.cell(row=row_num, column=6, value=nota.get_tipo_bodega_display() if nota.tipo_bodega else "")
        ws.cell(row=row_num, column=7, value=float(nota.neto or 0))
        ws.cell(row=row_num, column=8, value=float(nota.iva or 0))
        ws.cell(row=row_num, column=9, value=float(nota.total or 0))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="notas_venta_seleccionadas.xlsx"'
    wb.save(response)
    return response


@login_required
def status_pedido(request):
    if request.method == 'POST':
        accion = request.POST.get('accion', '').strip()

        if accion == 'excel':
            ids = _obtener_ids_seleccionados(request)
            if ids:
                notas_seleccionadas = NotaVenta.objects.filter(id__in=ids).select_related('vendedor').order_by('-fecha')
            else:
                notas_seleccionadas = NotaVenta.objects.select_related('vendedor').order_by('-fecha')
            return _exportar_notas_excel(notas_seleccionadas)

        if accion == 'eliminar':
            ids = _obtener_ids_seleccionados(request)
            if not ids:
                messages.error(request, 'Debes seleccionar al menos una nota de venta para eliminar.')
                return redirect('status_pedido')

            notas_seleccionadas = NotaVenta.objects.filter(id__in=ids).select_related('vendedor').order_by('-fecha')
            cantidad = notas_seleccionadas.count()
            notas_seleccionadas.delete()
            messages.success(request, f'Se eliminaron {cantidad} nota(s) de venta.')
            return redirect('status_pedido')

        messages.error(request, 'Acción no válida.')
        return redirect('status_pedido')

    notas_qs = NotaVenta.objects.select_related('vendedor').order_by('-fecha')
    paginator = Paginator(notas_qs, 10)
    page_number = request.GET.get('page')
    notas = paginator.get_page(page_number)
    return render(request, 'pedidos/status_pedido.html', {'notas': notas})


def _buscar_ubicaciones_central_para_detalle(detalle):
    codigo = (detalle.codigo or '').strip()
    ean = (detalle.ean or '').strip()
    dun = (detalle.dun or '').strip()

    if codigo:
        qs = Central.objects.filter(cod_sistema__iexact=codigo).order_by('id')
        if qs.exists():
            return qs

    if ean:
        qs = Central.objects.filter(cod_ean__iexact=ean).order_by('id')
        if qs.exists():
            return qs

    if dun:
        qs = Central.objects.filter(cod_dun__iexact=dun).order_by('id')
        if qs.exists():
            return qs

    return Central.objects.none()


def _construir_plan_picking(nota, bloquear=False):
    detalles_plan = []
    pendientes = []

    for detalle in nota.detalles.all().order_by('id'):
        cantidad_requerida = int(detalle.cantidad_solicitada or 0)
        faltante = cantidad_requerida

        ubicaciones_qs = _buscar_ubicaciones_central_para_detalle(detalle)
        if bloquear:
            ubicaciones_qs = ubicaciones_qs.select_for_update()

        asignaciones = []
        for item in ubicaciones_qs:
            cajas_disponibles = int(item.cajas or 0)
            if cajas_disponibles <= 0:
                continue

            extraer = min(faltante, cajas_disponibles)
            if extraer <= 0:
                continue

            asignaciones.append(
                {
                    'central_id': item.id,
                    'ubicacion': item.ubicacion or '-',
                    'cod_sistema': item.cod_sistema or '-',
                    'descripcion': item.descripcion or detalle.descripcion or '-',
                    'cajas_disponibles': cajas_disponibles,
                    'cajas_a_extraer': extraer,
                }
            )
            faltante -= extraer

            if faltante <= 0:
                break

        detalle_info = {
            'detalle': detalle,
            'cantidad_requerida': cantidad_requerida,
            'asignaciones': asignaciones,
            'faltante': max(faltante, 0),
        }
        detalles_plan.append(detalle_info)

        if faltante > 0:
            pendientes.append(
                f"{detalle.codigo or '-'} ({detalle.descripcion or '-'}) - faltan {faltante} cajas"
            )

    return detalles_plan, pendientes


@login_required
def crear_picking(request, nota_id):
    nota = get_object_or_404(
        NotaVenta.objects.select_related('vendedor').prefetch_related('detalles'),
        id=nota_id,
    )

    if request.method == 'POST':
        with transaction.atomic():
            nota_bloqueada = get_object_or_404(
                NotaVenta.objects.select_related('vendedor').prefetch_related('detalles'),
                id=nota_id,
            )
            plan, pendientes = _construir_plan_picking(nota_bloqueada, bloquear=True)

            if pendientes:
                messages.error(
                    request,
                    'No se pudo finalizar picking por stock insuficiente: ' + '; '.join(pendientes),
                )
                return redirect('crear_picking', nota_id=nota_id)

            for detalle_plan in plan:
                for asignacion in detalle_plan['asignaciones']:
                    central = Central.objects.get(id=asignacion['central_id'])
                    cajas_actuales = int(central.cajas or 0)
                    nuevas_cajas = max(cajas_actuales - int(asignacion['cajas_a_extraer']), 0)
                    central.cajas = nuevas_cajas
                    central.save(update_fields=['cajas'])

        messages.success(request, f"Picking finalizado para la Nota #{nota_id}.")
        return redirect('status_pedido')

    plan, pendientes = _construir_plan_picking(nota)

    return render(
        request,
        'pedidos/crear_picking.html',
        {
            'nota': nota,
            'plan_picking': plan,
            'pendientes': pendientes,
        },
    )


@login_required
@login_required
def descargar_picking_pdf(request, nota_id):
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
    except ImportError as e:
        logger.error(f'Error importando reportlab: {str(e)}')
        return HttpResponse(
            'Error: reportlab no está instalado en el servidor.',
            content_type='text/plain',
            status=500
        )

    try:
        nota = get_object_or_404(
            NotaVenta.objects.select_related('vendedor').prefetch_related('detalles'),
            id=nota_id,
        )
    except Exception as e:
        logger.error(f'Error obteniendo nota {nota_id}: {str(e)}')
        return HttpResponse(
            'Error: No se encontró la nota de venta.',
            content_type='text/plain',
            status=404
        )

    try:
        plan, _ = _construir_plan_picking(nota)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title = Paragraph(f"<b>Picking Nota de Venta #{nota.id}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Datos básicos
        basic_data = [
            ['Cliente:', nota.cliente or '-', 'Vendedor:', nota.vendedor.username],
            ['Fecha:', nota.fecha.strftime('%d/%m/%Y %H:%M') if nota.fecha else '-', '', '']
        ]
        basic_table = Table(basic_data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 2*inch])
        basic_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(basic_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Sección Cliente
        elements.append(Paragraph("<b>Datos del Cliente</b>", styles['Heading3']))
        cliente_data = [
            ['RUT:', nota.rut_cliente or '-', 'Teléfono:', nota.telefono or '-'],
            ['Giro:', nota.giro or '-', '', '']
        ]
        cliente_table = Table(cliente_data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 2*inch])
        cliente_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(cliente_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Dirección
        elements.append(Paragraph("<b>Dirección de Entrega</b>", styles['Heading3']))
        direccion_data = [
            ['Dirección:', nota.direccion or '-'],
            ['Comuna:', nota.comuna or '-'],
            ['Ciudad:', nota.ciudad or '-']
        ]
        dir_table = Table(direccion_data, colWidths=[1.2*inch, 5.5*inch])
        dir_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(dir_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Tabla de productos
        elements.append(Paragraph("<b>Ubicaciones asignadas para extracción</b>", styles['Heading3']))
        
        table_data = [['Cod.Sistema', 'Cod EAN', 'Cod DUN', 'Descripción', 'Cant.', 'Ubicación', 'Disp.', 'Extraer']]
        
        for item in plan:
            if item['asignaciones']:
                for asig in item['asignaciones']:
                    table_data.append([
                        item['detalle'].codigo or '-',
                        item['detalle'].ean or '-',
                        item['detalle'].dun or '-',
                        (item['detalle'].descripcion or '-')[:40],
                        str(item['cantidad_requerida']),
                        asig['ubicacion'],
                        str(asig['cajas_disponibles']),
                        str(asig['cajas_a_extraer'])
                    ])
            else:
                table_data.append([
                    item['detalle'].codigo or '-',
                    item['detalle'].ean or '-',
                    item['detalle'].dun or '-',
                    (item['detalle'].descripcion or '-')[:40],
                    str(item['cantidad_requerida']),
                    'Sin ubicación',
                    '0',
                    '0'
                ])

        productos_table = Table(table_data, colWidths=[0.9*inch, 1.1*inch, 1.1*inch, 2.2*inch, 0.5*inch, 1.1*inch, 0.5*inch, 0.6*inch])
        productos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(productos_table)

        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="picking_nota_{nota_id}.pdf"'
        response['Content-Type'] = 'application/pdf'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
        
    except Exception as e:
        logger.error(f'Error generando PDF para nota {nota_id}: {str(e)}', exc_info=True)
        return HttpResponse(
            f'Error al generar el PDF: {str(e)}',
            content_type='text/plain',
            status=500
        )


@login_required
def editar_nota(request, nota_id):
    nota = get_object_or_404(NotaVenta, id=nota_id)

    if request.method == 'POST':
        form = NotaVentaForm(request.POST, instance=nota)
        formset = DetalleNotaFormSet(request.POST, instance=nota, prefix='detalles')

        if form.is_valid() and formset.is_valid():
            nota_actualizada = form.save()
            detalles = formset.save(commit=False)

            for detalle_eliminado in formset.deleted_objects:
                detalle_eliminado.delete()

            for detalle in detalles:
                if detalle.precio_unitario and (not detalle.precio_x_caja or detalle.precio_x_caja != detalle.precio_unitario):
                    detalle.precio_x_caja = detalle.precio_unitario
                detalle.nota = nota_actualizada
                detalle.save()

            nota_actualizada.calcular_totales()
            messages.success(request, f"Nota #{nota_actualizada.id} actualizada correctamente.")
            return redirect('status_pedido')
    else:
        form = NotaVentaForm(instance=nota)
        formset = DetalleNotaFormSet(instance=nota, prefix='detalles')

    return render(
        request,
        'pedidos/editar_nota.html',
        {
            'nota': nota,
            'form': form,
            'formset': formset,
        }
    )


@login_required
def ver_nota_venta(request, nota_id):
    nota = get_object_or_404(
        NotaVenta.objects.select_related('vendedor').prefetch_related('detalles'),
        id=nota_id
    )

    detalles = nota.detalles.all().order_by('id')

    for detalle in detalles:
        total_linea = (Decimal(detalle.cantidad_solicitada) * detalle.precio_unitario).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        if detalle.importe_a_facturar != total_linea:
            detalle.importe_a_facturar = total_linea
            detalle.save(update_fields=['importe_a_facturar'])

    neto_calculado = sum(
        (Decimal(d.cantidad_solicitada) * d.precio_unitario for d in detalles),
        Decimal("0.00")
    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    iva_calculado = (neto_calculado * nota.IVA_PORCENTAJE).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )

    total_general = (neto_calculado + iva_calculado).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )

    if nota.neto != neto_calculado or nota.iva != iva_calculado or nota.total != total_general:
        nota.neto = neto_calculado
        nota.iva = iva_calculado
        nota.total = total_general
        nota.save(update_fields=['neto', 'iva', 'total'])

    return render(
        request,
        'pedidos/ver_nota_venta.html',
        {
            'nota': nota,
            'detalles': detalles,
            'total_general': total_general,
        }
    )


@login_required
@require_GET
def buscar_producto_por_codigo(request):
    codigo = request.GET.get("cod_sistema", "").strip()

    if not codigo:
        return _json_no_cache(
            {
                "found": False,
                "producto": {},
                "stocks": {
                    "bsf": 0,
                    "central": 0,
                    "total": 0,
                },
            }
        )

    bsf_base = Bsf.objects.exclude(cod_sistema__isnull=True).exclude(cod_sistema="")
    central_base = Central.objects.exclude(cod_sistema__isnull=True).exclude(cod_sistema="")

    bsf_exact = bsf_base.filter(cod_sistema__iexact=codigo)
    central_exact = central_base.filter(cod_sistema__iexact=codigo)

    bsf_qs = bsf_exact if bsf_exact.exists() else bsf_base.filter(cod_sistema__icontains=codigo)
    central_qs = central_exact if central_exact.exists() else central_base.filter(cod_sistema__icontains=codigo)

    stock_bsf = bsf_qs.aggregate(total=Coalesce(Sum("cajas"), 0))["total"] or 0
    stock_central = central_qs.aggregate(total=Coalesce(Sum("cajas"), 0))["total"] or 0
    stock_total = stock_bsf + stock_central

    producto = bsf_qs.first() or central_qs.first()

    if not producto:
        return _json_no_cache(
            {
                "found": False,
                "producto": {},
                "stocks": {
                    "bsf": 0,
                    "central": 0,
                    "total": 0,
                },
            }
        )

    return _json_no_cache(
        {
            "found": True,
            "producto": {
                "cod_sistema": producto.cod_sistema or "",
                "descripcion": producto.descripcion or "",
                "cod_ean": producto.cod_ean or "",
                "cod_dun": producto.cod_dun or "",
            },
            "stocks": {
                "bsf": stock_bsf,
                "central": stock_central,
                "total": stock_total,
            },
        }
    )
