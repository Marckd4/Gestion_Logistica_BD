
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from functools import wraps

from bodegabsf.forms import BsfForm
from .models import Bsf  # importar el modelo
from pedidos.models import UserModulePermission


def _user_has_module_permission(user, module, action):
    if user.is_superuser:
        return True

    permiso = UserModulePermission.objects.filter(user=user, module=module).first()
    if not permiso:
        return False

    permisos = {
        'view': permiso.can_view,
        'create': permiso.can_create,
        'edit': permiso.can_edit,
        'delete': permiso.can_delete,
        'export': permiso.can_export,
        'report': permiso.can_report,
    }
    return permisos.get(action, False)


def require_module_permission(module, action):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if _user_has_module_permission(request.user, module, action):
                return view_func(request, *args, **kwargs)

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'ok': False, 'message': 'No tienes permisos para esta acción.'}, status=403)

            messages.error(request, 'No tienes permisos para esta acción en este módulo.')
            return redirect('inicio')

        return _wrapped_view

    return decorator



@login_required
@require_module_permission('bodegabsf', 'view')
def data(request):
    bsfs = Bsf.objects.all()
    return render(request, 'index.html', context={'bsfs': bsfs})

# area formulario 

def formulario(request):
    if request.method == 'POST':
        form = BsfForm(request.POST)
        if form.is_valid():
             form.save()
             return HttpResponseRedirect('/bodegabsf')
    else:
        form = BsfForm()
        
        
    return render( request, 'bsf_form.html',{'form': form})



# area de exportar excel 
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import Bsf

def limpiar(valor):
    if valor is None:
        return ""
    return str(valor)

@login_required
@require_module_permission('bodegabsf', 'export')
def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    columnas = [
        "Categoria", "Empresa", "Ubicacion", "Cod_Ean", "Cod_Dun", "Cod_Sistema",
        "Descripcion", "Unidad", "Pack", "FactorX", "Cajas", "Saldo",
        "Stock_Fisico", "Observacion", "Fecha_Venc", "Fecha_Imp",
        "Fecha_Inv", "Encargado"
    ]

    # Encabezados
    for col_num, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=titulo)
        cell.font = Font(bold=True)

    data = Bsf.objects.all()

    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=limpiar(item.categoria))
        ws.cell(row=row_num, column=2, value=limpiar(item.empresa))
        ws.cell(row=row_num, column=3, value=limpiar(item.ubicacion))
        ws.cell(row=row_num, column=4, value=limpiar(item.cod_ean))
        ws.cell(row=row_num, column=5, value=limpiar(item.cod_dun))
        ws.cell(row=row_num, column=6, value=limpiar(item.cod_sistema))
        ws.cell(row=row_num, column=7, value=limpiar(item.descripcion))
        ws.cell(row=row_num, column=8, value=limpiar(item.unidad))
        ws.cell(row=row_num, column=9, value=limpiar(item.pack))
        ws.cell(row=row_num, column=10, value=limpiar(item.factorx))
        ws.cell(row=row_num, column=11, value=limpiar(item.cajas))
        ws.cell(row=row_num, column=12, value=limpiar(item.saldo))
        ws.cell(row=row_num, column=13, value=limpiar(item.stock_fisico))
        ws.cell(row=row_num, column=14, value=limpiar(item.observacion))
        ws.cell(row=row_num, column=15, value=limpiar(item.fecha_venc))
        ws.cell(row=row_num, column=16, value=limpiar(item.fecha_imp))
        ws.cell(row=row_num, column=17, value=limpiar(item.fecha_inv))
        ws.cell(row=row_num, column=18, value=limpiar(item.encargado))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=\"inventario_bodega_BSF.xlsx\"'

    wb.save(response)
    return response


# eliminar y editar tabla 


from django.shortcuts import render, redirect, get_object_or_404
from .models import Bsf
from .forms import BsfForm

# ---------- EDITAR ----------
@login_required
@require_module_permission('bodegabsf', 'edit')
def editar_bsf(request, id):
    bsf = get_object_or_404(Bsf, id=id)

    if request.method == "POST":
        form = BsfForm(request.POST, instance=bsf)
        if form.is_valid():
            form.save()
            return redirect('data')  # ← Ajusta al nombre de tu vista principal
    else:
        form = BsfForm(instance=bsf)

    return render(request, 'editar_bsf.html', {'form': form, 'bsf': bsf})


# ---------- ELIMINAR ----------
@login_required
@require_module_permission('bodegabsf', 'delete')
def eliminar_bsf(request, id):
    bsf = get_object_or_404(Bsf, id=id)

    if request.method == "POST":
        bsf.delete()
        return redirect('data')   # ← Ajusta al nombre de tu vista principal

    return render(request, 'eliminar_bsf.html', {'bsf': bsf})



# autocompletar del dun en formulario 

from django.shortcuts import render
from django.http import JsonResponse
from .models import Bsf
from .forms import BsfForm

@login_required
@require_module_permission('bodegabsf', 'create')
def formulario(request, id=None):
    mensaje = ""
    if id:  # modo edición
        instancia = Bsf.objects.get(id=id)
    else:
        instancia = None

    if request.method == "POST":
        form = BsfForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            return redirect('/bodegabsf')
            mensaje = "✅ Producto guardado correctamente."
    else:
        form = BsfForm(instance=instancia)

    return render(request, 'bsf_form.html', {"form": form, "mensaje": mensaje})

@login_required
@require_module_permission('bodegabsf', 'view')
def buscar_producto(request):
    cod_dun = request.GET.get('cod_dun', '').strip()
    resultados = []

    if cod_dun:
        productos = Bsf.objects.filter(cod_dun__icontains=cod_dun)[:10]  # máximo 10 coincidencias
        for p in productos:
            resultados.append({
                "cod_dun": p.cod_dun,
                "cod_ean": p.cod_ean,
                "cod_sistema": p.cod_sistema,
                "descripcion": p.descripcion,
            })

    return JsonResponse({"resultados": resultados})


@login_required
@require_module_permission('bodegabsf', 'view')
@ensure_csrf_cookie
def cambio_ubicacion(request):
    return render(request, 'bodegabsf/cambio_ubicacion.html')


@login_required
@require_module_permission('bodegabsf', 'view')
@require_GET
def buscar_por_ubicacion(request):
    ubicacion = request.GET.get('ubicacion', '').strip()
    if not ubicacion:
        return JsonResponse({'ok': False, 'message': 'Debes ingresar una ubicacion.'}, status=400)

    productos = Bsf.objects.filter(ubicacion__iexact=ubicacion)
    if not productos.exists():
        return JsonResponse({'ok': False, 'message': 'No se encontraron productos en esa ubicacion.'}, status=404)

    lista_productos = []
    for p in productos:
        lista_productos.append({
            'id': p.id,
            'cod_dun': p.cod_dun or '',
            'cod_ean': p.cod_ean or '',
            'cod_sistema': p.cod_sistema or '',
            'descripcion': p.descripcion or '',
            'cajas': p.cajas if p.cajas is not None else 0,
            'ubicacion': p.ubicacion or '',
        })

    return JsonResponse({
        'ok': True,
        'productos': lista_productos
    })


@login_required
@require_module_permission('bodegabsf', 'edit')
@require_POST
def mover_producto(request):
    from django.db import transaction
    
    producto_id = request.POST.get('producto_id', '').strip()
    cantidad_mover = request.POST.get('cantidad_mover', '').strip()
    nueva_ubicacion = request.POST.get('nueva_ubicacion', '').strip()

    if not producto_id:
        return JsonResponse({'ok': False, 'message': 'Producto invalido.'}, status=400)

    if not cantidad_mover or not cantidad_mover.isdigit():
        return JsonResponse({'ok': False, 'message': 'Cantidad invalida.'}, status=400)
    
    cantidad_mover = int(cantidad_mover)
    if cantidad_mover <= 0:
        return JsonResponse({'ok': False, 'message': 'La cantidad debe ser mayor a 0.'}, status=400)

    if not nueva_ubicacion:
        return JsonResponse({'ok': False, 'message': 'Debes ingresar la nueva ubicacion.'}, status=400)

    producto_origen = Bsf.objects.filter(id=producto_id).first()
    if not producto_origen:
        return JsonResponse({'ok': False, 'message': 'Producto no encontrado.'}, status=404)

    if producto_origen.cajas is None or producto_origen.cajas < cantidad_mover:
        return JsonResponse({'ok': False, 'message': f'No hay suficientes cajas. Disponible: {producto_origen.cajas or 0}'}, status=400)

    try:
        with transaction.atomic():
            # Restar cajas del origen
            producto_origen.cajas -= cantidad_mover
            if producto_origen.stock_fisico:
                producto_origen.stock_fisico = producto_origen.cajas * (producto_origen.factorx or 1)
            
            if producto_origen.cajas == 0:
                producto_origen.delete()
            else:
                producto_origen.save()

            # Buscar si existe el mismo producto en la nueva ubicacion
            producto_destino = Bsf.objects.filter(
                cod_sistema=producto_origen.cod_sistema,
                ubicacion__iexact=nueva_ubicacion
            ).first()

            if producto_destino:
                # Sumar cajas al destino existente
                producto_destino.cajas = (producto_destino.cajas or 0) + cantidad_mover
                if producto_destino.stock_fisico is not None:
                    producto_destino.stock_fisico = producto_destino.cajas * (producto_destino.factorx or 1)
                producto_destino.save()
            else:
                # Crear nuevo registro en la nueva ubicacion
                Bsf.objects.create(
                    categoria=producto_origen.categoria,
                    empresa=producto_origen.empresa,
                    ubicacion=nueva_ubicacion,
                    cod_ean=producto_origen.cod_ean,
                    cod_dun=producto_origen.cod_dun,
                    cod_sistema=producto_origen.cod_sistema,
                    descripcion=producto_origen.descripcion,
                    unidad=producto_origen.unidad,
                    pack=producto_origen.pack,
                    factorx=producto_origen.factorx,
                    cajas=cantidad_mover,
                    saldo=producto_origen.saldo,
                    stock_fisico=cantidad_mover * (producto_origen.factorx or 1) if producto_origen.factorx else None,
                    observacion=producto_origen.observacion,
                    fecha_inv=producto_origen.fecha_inv,
                    encargado=producto_origen.encargado,
                    fecha_venc=producto_origen.fecha_venc,
                    fecha_imp=producto_origen.fecha_imp,
                    numero_contenedor=producto_origen.numero_contenedor,
                    cant_solicitada=producto_origen.cant_solicitada,
                    pedido=producto_origen.pedido,
                )

        return JsonResponse({
            'ok': True,
            'message': f'Se movieron {cantidad_mover} cajas a {nueva_ubicacion} correctamente.',
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'message': f'Error al mover producto: {str(e)}'}, status=500)



# resumen de datos 

from django.shortcuts import render
from .models import Bsf

@login_required
@require_module_permission('bodegabsf', 'view')
def resumen_bsf(request):
    datos = Bsf.objects.all().values(
        "cod_dun",
        "cod_ean",
        "cod_sistema",
        "descripcion",
        "cajas",
        "stock_fisico",
        "ubicacion",
    )
    return render(request, "resumen_bsf.html", {"datos": list(datos)})


# views.py
from django.contrib.auth.decorators import login_required

@login_required
def crear_bsf(request):
    if request.method == 'POST':
        form = BsfForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_bsf')
    else:
        form = BsfForm()
    return render(request, 'bodegabsf/bsf_form.html', {'form': form})


# pedidos
from django.shortcuts import render
from django.db.models import Sum
from .models import Bsf

@login_required
@require_module_permission('bodegabsf', 'report')
def resumen_pedidos(request):
    fecha = request.GET.get('fecha')
    pedido = request.GET.get('pedido')
    ubicacion = request.GET.get('ubicacion')

    # 🔴 BASE: excluir pedidos vacíos o en 0
    queryset = Bsf.objects.exclude(pedido__isnull=True).exclude(pedido=0)

    if fecha:
        queryset = queryset.filter(fecha_inv=fecha)

    if pedido:
        queryset = queryset.filter(pedido=pedido)

    if ubicacion:
        queryset = queryset.filter(ubicacion__icontains=ubicacion)

    resumen = (
        queryset
        .values(
            'pedido',
            'ubicacion',
            'cod_ean',
            'cod_dun',
            'cod_sistema',
            'descripcion'
        )
        .annotate(
            total_cant_solicitada=Sum('cant_solicitada')
        )
        .order_by(
            'pedido',
            'ubicacion',
            'descripcion'
        )
    )

    return render(request, 'resumen_pedidos.html', {
        'resumen': resumen
    })
# excel pedidos
import openpyxl
from django.http import HttpResponse
from django.db.models import Sum
from .models import Bsf


@login_required
@require_module_permission('bodegabsf', 'export')
def resumen_pedidos_excel(request):
    fecha = request.GET.get('fecha')
    pedido = request.GET.get('pedido')
    ubicacion = request.GET.get('ubicacion')

    # 🔴 SOLO pedidos válidos
    queryset = Bsf.objects.filter(pedido__gt=0)

    if fecha:
        queryset = queryset.filter(fecha_inv=fecha)

    if pedido:
        queryset = queryset.filter(pedido=pedido)

    if ubicacion:
        queryset = queryset.filter(ubicacion__icontains=ubicacion)

    resumen = (
        queryset
        .values(
            'pedido',
            'ubicacion',
            'cod_ean',
            'cod_dun',
            'cod_sistema',
            'descripcion'
        )
        .annotate(
            total_cant_solicitada=Sum('cant_solicitada')
        )
        .order_by('pedido', 'ubicacion')
    )

    # =========================
    # CREAR EXCEL
    # =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Pedidos"

    ws.append([
        'Pedido',
        'Ubicación',
        'Cod EAN',
        'Cod DUN',
        'Cod Sistema',
        'Descripción',
        'Total Cant. Solicitada'
    ])

    for r in resumen:
        ws.append([
            r['pedido'],
            r['ubicacion'],
            r['cod_ean'],
            r['cod_dun'],
            r['cod_sistema'],
            r['descripcion'],
            r['total_cant_solicitada'] or 0
        ])

    # =========================
    # RESPUESTA HTTP (CLAVE)
    # =========================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=resumen_pedidos.xlsx'

    wb.save(response)

    return response  # ✅ ESTO SOLUCIONA EL ERROR


# solicitudes a facturacion 

from django.shortcuts import render
from django.db.models import Sum
from .models import Bsf


@login_required
@require_module_permission('bodegabsf', 'report')
def resumen_general_pedidos(request):
    fecha = request.GET.get('fecha')
    pedido = request.GET.get('pedido')
    cod_ean = request.GET.get('cod_ean')
    cod_dun = request.GET.get('cod_dun')
    cod_sistema = request.GET.get('cod_sistema')

    queryset = Bsf.objects.filter(pedido__gt=0)

    if fecha:
        queryset = queryset.filter(fecha_inv=fecha)

    if pedido:
        queryset = queryset.filter(pedido=pedido)

    if cod_ean:
        queryset = queryset.filter(cod_ean__icontains=cod_ean)

    if cod_dun:
        queryset = queryset.filter(cod_dun__icontains=cod_dun)

    if cod_sistema:
        queryset = queryset.filter(cod_sistema__icontains=cod_sistema)

    resumen = (
        queryset
        .values(
            'pedido',
            'cod_ean',
            'cod_dun',
            'cod_sistema',
            'descripcion'
        )
        .annotate(
            total_cant_solicitada=Sum('cant_solicitada')
        )
        .order_by('pedido', 'descripcion')
    )

    return render(request, 'resumen_general_pedidos.html', {
        'resumen': resumen
    })


# 02 
import openpyxl
from django.http import HttpResponse
from django.db.models import Sum
from .models import Bsf


@login_required
@require_module_permission('bodegabsf', 'export')
def resumen_general_pedidos_excel(request):
    fecha = request.GET.get('fecha')
    pedido = request.GET.get('pedido')
    cod_ean = request.GET.get('cod_ean')
    cod_dun = request.GET.get('cod_dun')
    cod_sistema = request.GET.get('cod_sistema')

    queryset = Bsf.objects.filter(pedido__gt=0)

    if fecha:
        queryset = queryset.filter(fecha_inv=fecha)

    if pedido:
        queryset = queryset.filter(pedido=pedido)

    if cod_ean:
        queryset = queryset.filter(cod_ean__icontains=cod_ean)

    if cod_dun:
        queryset = queryset.filter(cod_dun__icontains=cod_dun)

    if cod_sistema:
        queryset = queryset.filter(cod_sistema__icontains=cod_sistema)

    resumen = (
        queryset
        .values(
            'pedido',
            'cod_ean',
            'cod_dun',
            'cod_sistema',
            'descripcion'
        )
        .annotate(
            total_cant_solicitada=Sum('cant_solicitada')
        )
        .order_by('pedido')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen General Pedidos"

    ws.append([
        'Pedido',
        'Cod EAN',
        'Cod DUN',
        'Cod Sistema',
        'Descripción',
        'Total Cant. Solicitada'
    ])

    for r in resumen:
        ws.append([
            r['pedido'],
            r['cod_ean'],
            r['cod_dun'],
            r['cod_sistema'],
            r['descripcion'],
            r['total_cant_solicitada'] or 0
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename=resumen_general_pedidos.xlsx'
    )

    wb.save(response)
    return response



# importar
import pandas as pd
from datetime import date
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction

from .models import Bsf
from .forms import ImportarExcelForm


# -----------------------------
# LIMPIEZA DE DATOS
# -----------------------------
def limpiar(valor):
    """
    - NaN / vacío → None
    - string vacío → None
    - 0 → 0
    """
    if pd.isna(valor):
        return None
    if isinstance(valor, str):
        valor = valor.strip()
        return valor if valor else None
    return valor


def limpiar_fecha(valor):
    """
    Acepta solo fechas válidas
    Texto como 'SIN INFORMACION' → None
    """
    if pd.isna(valor):
        return None

    if isinstance(valor, pd.Timestamp):
        return valor.date()

    # Texto u otro tipo → None
    return None


# -----------------------------
# IMPORTAR EXCEL
# -----------------------------
@login_required
@require_module_permission('bodegabsf', 'create')
def importar_excel(request):
    if request.method == "POST":
        form = ImportarExcelForm(request.POST, request.FILES)

        if form.is_valid():
            archivo = request.FILES["archivo"]

            try:
                df = pd.read_excel(archivo)

                nuevos = []
                actualizar = []

                # Obtener códigos existentes
                cods = df["cod_ean"].dropna().unique()
                existentes = {
                    obj.cod_ean: obj
                    for obj in Bsf.objects.filter(cod_ean__in=cods)
                }

                for _, fila in df.iterrows():
                    cod_ean = limpiar(fila.get("cod_ean"))
                    if not cod_ean:
                        continue

                    data = {
                        "categoria": limpiar(fila.get("categoria")),
                        "empresa": limpiar(fila.get("empresa")),
                        "ubicacion": limpiar(fila.get("ubicacion")),
                        "cod_dun": limpiar(fila.get("cod_dun")),
                        "cod_sistema": limpiar(fila.get("cod_sistema")),
                        "descripcion": limpiar(fila.get("descripcion")),
                        "unidad": limpiar(fila.get("unidad")),
                        "pack": int(limpiar(fila.get("pack"))) if limpiar(fila.get("pack")) is not None else None,
                        "factorx": float(limpiar(fila.get("factorx"))) if limpiar(fila.get("factorx")) is not None else None,
                        "cajas": int(limpiar(fila.get("cajas"))) if limpiar(fila.get("cajas")) is not None else None,
                        "saldo": int(limpiar(fila.get("saldo"))) if limpiar(fila.get("saldo")) is not None else None,
                        "stock_fisico": int(limpiar(fila.get("stock_fisico"))) if limpiar(fila.get("stock_fisico")) is not None else None,
                        "observacion": limpiar(fila.get("observacion")),
                        "fecha_inv": limpiar_fecha(fila.get("fecha_inv")),
                        "encargado": limpiar(fila.get("encargado")),
                        "fecha_venc": limpiar_fecha(fila.get("fecha_venc")),
                        "fecha_imp": date.today(),
                        "numero_contenedor": limpiar(fila.get("numero_contenedor")),
                        "cant_solicitada": int(limpiar(fila.get("cant_solicitada"))) if limpiar(fila.get("cant_solicitada")) is not None else None,
                        "pedido": int(limpiar(fila.get("pedido"))) if limpiar(fila.get("pedido")) is not None else None,
                    }

                    if cod_ean in existentes:
                        obj = existentes[cod_ean]
                        for campo, valor in data.items():
                            setattr(obj, campo, valor)
                        actualizar.append(obj)
                    else:
                        nuevos.append(Bsf(cod_ean=cod_ean, **data))

                # Guardado optimizado
                with transaction.atomic():
                    if nuevos:
                        Bsf.objects.bulk_create(nuevos, batch_size=500)
                    if actualizar:
                        Bsf.objects.bulk_update(
                            actualizar,
                            fields=data.keys(),
                            batch_size=500
                        )

                messages.success(
                    request,
                    f"Importación OK: {len(nuevos)} nuevos / {len(actualizar)} actualizados"
                )
                return redirect("importar_excel")

            except Exception as e:
                messages.error(request, f"Error al importar: {e}")

    else:
        form = ImportarExcelForm()

    return render(request, "importar_excel.html", {"form": form})


from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect


def es_admin(user):
    return user.is_superuser


@login_required
@user_passes_test(es_admin)
@require_module_permission('bodegabsf', 'delete')
def borrar_todo_bsf(request):
    if request.method == "POST":
        with transaction.atomic():
            total = Bsf.objects.count()
            Bsf.objects.all().delete()

        messages.success(
            request,
            f"Se eliminaron {total} registros correctamente."
        )

    return redirect("importar_excel")

